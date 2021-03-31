# Coded by MercyMist
# Project: Student Registration Form (GUI and Openpyxl)

# import openpyxl and tkinter

from openpyxl import *
from tkinter import *

from openpyxl import Workbook
wb = Workbook()
filepath = 'E:\Student Registration.xlsx'
sheet = wb.active
wb.save('E:\Student Registration.xlsx')

def excel():

    # resize width of column in spreadsheet

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data

    sheet.cell(row = 1, column = 1).value = 'Name'
    sheet.cell(row = 1, column = 2).value = 'Course'
    sheet.cell(row = 1, column = 3).value = 'Semester'
    sheet.cell(row = 1, column = 4).value = 'Form Number'
    sheet.cell(row = 1, column = 5).value = 'Contact Number'
    sheet.cell(row = 1, column = 6).value = 'E-mail ID'
    sheet.cell(row = 1, column = 7).value = 'Address'

    wb.save('E:\Student Registration.xlsx')

# Function to set focus (cursor)

def focus1(event):
    # set focus on the course field box
    course_field.focus_set()

def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus_set()

def focus6(event):
    address_field.focus_set()

# Function for clearing the contents of text entry boxes

def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

# Function to take data from GUI window and write to excel file

def insert():
    
    # print "empty input" if no data is filled

    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):

        print("Empty Input")

    else:

        # assigning max row and max column value upto
        # which data is written in an excel to variable

        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text as string which we
        # write into excel sheet at particular location

        sheet.cell(row = current_row + 1, column = 1).value = name_field.get()
        sheet.cell(row = current_row + 1, column = 2).value = course_field.get()
        sheet.cell(row = current_row + 1, column = 3).value = sem_field.get()
        sheet.cell(row = current_row + 1, column = 4).value = form_no_field.get()
        sheet.cell(row = current_row + 1, column = 5).value = contact_no_field.get()
        sheet.cell(row = current_row + 1, column = 6).value = email_id_field.get()
        sheet.cell(row = current_row + 1, column = 7).value = address_field.get()

        # save the file

        wb.save('E:\Student Registration.xlsx')

        # set focus on the first field

        name_field.focus_set()

        # call the clear() function

        clear()

# Driver Code

if __name__ == "__main__":

    # create a GUI window

    root = Tk()

    # background color and Title

    root.configure(background = 'light green')

    root.title('Student Registration Form | MercyMist')

    root.geometry("500x300")

    excel()

    # creating Labels for Form

    heading = Label(root, text = "Form", bg = "light green")

    name = Label(root, text = "Name", bg = "light green")

    course = Label(root, text = "Course", bg = "light green")

    sem = Label(root, text = "Semester", bg = "light green")

    form_no = Label(root, text = "Form Number", bg = "light green")

    contact_no = Label(root, text = "Contact Number", bg = "light green")

    email_id = Label(root, text = "E-mail ID", bg = "light green")

    address = Label(root, text = "Address", bg = "light green")

    # grid method is used for placing widgets in
    # table like structure

    heading.grid(row = 0, column = 1)
    name.grid(row = 1, column = 0)
    course.grid(row = 2, column = 0)
    sem.grid(row = 3, column = 0)
    form_no.grid(row = 4, column = 0)
    contact_no.grid(row = 5, column = 0)
    email_id.grid(row = 6, column = 0)
    address.grid(row = 7, column = 0)

    # create text entry box for entering information

    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    # bind method is used to bind functions with events

    # whenever enter key is pressed then call focus1 function

    name_field.bind('<Return>', focus1)

    course_field.bind('<Return>', focus2)

    sem_field.bind('<Return>', focus3)

    form_no_field.bind('<Return>', focus4)

    contact_no_field.bind('<Return>', focus5)

    email_id_field.bind('<Return>', focus6)

    # grid method is used for placing widgets in
    # table like structure

    name_field.grid(row = 1, column = 1, ipadx = '100')
    course_field.grid(row = 2, column = 1, ipadx = '100')
    sem_field.grid(row = 3, column = 1, ipadx = '100')
    form_no_field.grid(row = 4, column = 1, ipadx = '100')
    contact_no_field.grid(row = 5, column = 1, ipadx = '100')
    email_id_field.grid(row = 6, column = 1, ipadx = '100')
    address_field.grid(row = 7, column = 1, ipadx = '100')

    # call excel function

    excel()

    # create submit button and place into the root window

    submit = Button(root, text = 'Submit', fg = 'Black', bg = 'Orange', command = insert)

    submit.grid(row = 8, column = 1)

    # start the GUI

    root.mainloop()

    

    
    
